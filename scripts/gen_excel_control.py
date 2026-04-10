"""
Genera el Excel de Control de Automatizacion de APIs de Policysense.
Fuente de datos: docs/api-inventory.md
Salida: docs/API_Control_Automatizacion.xlsx
"""
import subprocess, sys
subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "-q"], check=True)

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import OrderedDict

wb = Workbook()

# ============================================================
# DATA
# ============================================================
endpoints = []
id_counter = {}

PREFIX_MAP = {
    'Policy Admin': 'PA', 'Facturacion y Cobranza': 'FC', 'Siniestros': 'SI',
    'Distribution Management': 'DM', 'Ventas': 'VE', 'Reaseguro': 'RE',
    'Productos': 'PR', 'Admin Usuarios': 'AU', 'Servicio de Campo (SMC)': 'SM',
    'Portal Autoservicio': 'PO', 'Configuracion': 'CF', 'OAuth (Transversal)': 'OA',
}

def add(modulo, microservicio, path_base, metodo, endpoint, descripcion, prioridad, automatizado, tg, sla):
    prefix = PREFIX_MAP.get(modulo, 'XX')
    id_counter[prefix] = id_counter.get(prefix, 0) + 1
    eid = f'POL-{prefix}{id_counter[prefix]:03d}'
    endpoints.append({
        'id': eid, 'modulo': modulo, 'microservicio': microservicio,
        'path_base': path_base, 'metodo': metodo, 'endpoint': endpoint,
        'descripcion': descripcion, 'prioridad': prioridad,
        'automatizado': automatizado, 'tg': tg, 'sla': sla,
        'estado': 'Completado' if automatizado == 'Si' else 'Pendiente',
        'notas': ''
    })

# --- Policy Admin - RatingService ---
s, p = 'RatingService', '/rest/ratingservice/v1'
add('Policy Admin', s, p, 'POST', '/Risk/BasicAnnualPremiumForAllCoverage', 'Calculo de primas anuales por cobertura', 'ALTA', 'No', '--', 2000)
add('Policy Admin', s, p, 'POST', '/Risk/GrossAllModalPremiumForAllCoverages', 'Calculo importes a facturar por frecuencia', 'ALTA', 'No', '--', 2000)
add('Policy Admin', s, p, 'POST', '/Risk/GrossModalPremiumForSelectedCoverages', 'Importes a facturar coberturas seleccionadas', 'ALTA', 'No', '--', 2000)
add('Policy Admin', s, p, 'POST', '/Risk/BasicAnnualPremiumAllCoverageSeveralRisks', 'Primas anuales para multiples riesgos', 'ALTA', 'No', '--', 2000)
add('Policy Admin', s, p, 'GET', '/Risk/PolicyInformation', 'Info de poliza (OfficialPolicyNumber, EffectiveDate)', 'MEDIA', 'No', '--', 2000)
add('Policy Admin', s, p, 'GET', '/RiskString', 'Procesar cadena de riesgo', 'BAJA', 'No', '--', 2000)

# --- Policy Admin - PolicyAdminService ---
s, p = 'PolicyAdminService', '/rest/policyadmin/v1'
add('Policy Admin', s, p, 'GET', '/quotation/all', 'Listar todas las cotizaciones (paginado)', 'ALTA', 'Si', 'TG01', 2000)
add('Policy Admin', s, p, 'DELETE', '/quotation/all', 'Eliminar todas las cotizaciones (key)', 'BAJA', 'No', '--', 2000)
add('Policy Admin', s, p, 'GET', '/quotation', 'Obtener cotizacion por QuotationId', 'ALTA', 'Si', 'TG01', 2000)
add('Policy Admin', s, p, 'POST', '/quotation', 'Crear cotizacion', 'ALTA', 'Si', 'TG01', 2000)
add('Policy Admin', s, p, 'PUT', '/quotation', 'Actualizar cotizacion', 'ALTA', 'No', '--', 2000)
add('Policy Admin', s, p, 'DELETE', '/quotation/delete', 'Eliminar cotizacion', 'MEDIA', 'No', '--', 2000)
add('Policy Admin', s, p, 'GET', '/quotation/newid', 'Generar nuevo ID de cotizacion', 'MEDIA', 'No', '--', 2000)
add('Policy Admin', s, p, 'GET', '/quotation/allanonynousbyquotation', 'Cotizaciones anonimas por criterio', 'MEDIA', 'No', '--', 2000)
add('Policy Admin', s, p, 'GET', '/quotation/allanonymousbyclient', 'Cotizaciones anonimas por cliente', 'MEDIA', 'No', '--', 2000)
add('Policy Admin', s, p, 'GET', '/quotation/allbyclient', 'Cotizaciones por cliente (ClientCode, Email)', 'ALTA', 'No', '--', 2000)
add('Policy Admin', s, p, 'GET', '/quotation/numberofquotations', 'Contar cotizaciones por cliente', 'MEDIA', 'No', '--', 2000)
add('Policy Admin', s, p, 'GET', '/quotation/allbyclientgroupbyproduct', 'Cotizaciones agrupadas por producto', 'MEDIA', 'No', '--', 2000)
add('Policy Admin', s, p, 'GET', '/quotation/orphans/count', 'Contar cotizaciones huerfanas', 'BAJA', 'No', '--', 2000)
add('Policy Admin', s, p, 'DELETE', '/quotation/orphans', 'Eliminar cotizaciones huerfanas', 'BAJA', 'No', '--', 2000)
add('Policy Admin', s, p, 'POST', '/quotation/life', 'Crear cotizacion vida', 'ALTA', 'No', '--', 2000)
add('Policy Admin', s, p, 'PUT', '/quotation/life', 'Actualizar cotizacion vida', 'ALTA', 'No', '--', 2000)
add('Policy Admin', s, p, 'GET', '/quotation/life/{QuotationId}', 'Obtener cotizacion vida', 'ALTA', 'No', '--', 2000)
add('Policy Admin', s, p, 'GET', '/quotation/broker', 'Cotizaciones/solicitudes/polizas por broker', 'ALTA', 'No', '--', 2000)
add('Policy Admin', s, p, 'GET', '/quotation/quotationsource', 'Cotizaciones con documentos fuente', 'MEDIA', 'No', '--', 2000)
add('Policy Admin', s, p, 'GET', '/quotationclient', 'Obtener cliente de cotizacion', 'MEDIA', 'No', '--', 2000)
add('Policy Admin', s, p, 'GET', '/quotationcoverage', 'Coberturas de una cotizacion', 'ALTA', 'No', '--', 2000)
add('Policy Admin', s, p, 'GET', '/quotationquantity', 'Total de cotizaciones', 'MEDIA', 'No', '--', 2000)
add('Policy Admin', s, p, 'GET', '/quotationquantity/ByDistributionChannel', 'Cotizaciones por canal de distribucion', 'MEDIA', 'No', '--', 2000)
add('Policy Admin', s, p, 'POST', '/quotationpdf', 'Generar PDF de cotizacion', 'MEDIA', 'No', '--', 2000)
add('Policy Admin', s, p, 'GET', '/quotationpdf', 'Obtener PDF de cotizacion', 'MEDIA', 'No', '--', 2000)
add('Policy Admin', s, p, 'DELETE', '/quotationpdf', 'Eliminar PDF', 'BAJA', 'No', '--', 2000)
add('Policy Admin', s, p, 'GET', '/quotationpdf/quotationlifepdf/{OfficialQuotationId}', 'PDF de cotizacion vida', 'MEDIA', 'No', '--', 2000)
add('Policy Admin', s, p, 'GET', '/quotationclientaddress', 'Direcciones del cliente', 'MEDIA', 'No', '--', 2000)
add('Policy Admin', s, p, 'POST', '/quotationclientaddress/save', 'Guardar direccion', 'MEDIA', 'No', '--', 2000)
add('Policy Admin', s, p, 'PUT', '/quotationclientaddress/update', 'Actualizar direccion', 'MEDIA', 'No', '--', 2000)
add('Policy Admin', s, p, 'GET', '/quotationclientaddress/GetByQuotation', 'Direcciones por cotizacion', 'MEDIA', 'No', '--', 2000)
add('Policy Admin', s, p, 'POST', '/insuranceduration', 'Crear duracion de seguro', 'MEDIA', 'No', '--', 2000)
add('Policy Admin', s, p, 'GET', '/product', 'Productos por canal de distribucion', 'ALTA', 'No', '--', 2000)
add('Policy Admin', s, p, 'GET', '/applicationformat', 'Formatos de solicitud por producto', 'MEDIA', 'No', '--', 2000)
add('Policy Admin', s, p, 'GET', '/applicationformat/GetForClientQuotation', 'Formato de solicitud cliente', 'MEDIA', 'No', '--', 2000)

# --- Policy Admin - UnderwritingService ---
s, p = 'UnderwritingService', '/rest/underwritingservice/v1'
add('Policy Admin', s, p, 'GET', '/requirement/getbyclient', 'Requerimientos por cliente', 'ALTA', 'No', '--', 2000)
add('Policy Admin', s, p, 'POST', '/requirement', 'Crear requerimiento de suscripcion', 'ALTA', 'No', '--', 2000)
add('Policy Admin', s, p, 'GET', '/requirement', 'Obtener requerimiento por RecGUID', 'MEDIA', 'No', '--', 2000)
add('Policy Admin', s, p, 'GET', '/requirement/getbycase', 'Requerimientos por caso', 'MEDIA', 'No', '--', 2000)
add('Policy Admin', s, p, 'GET', '/requirement/getbyquotation', 'Requerimientos por cotizacion', 'MEDIA', 'No', '--', 2000)
add('Policy Admin', s, p, 'GET', '/requirement/getpendingbychannel', 'Requerimientos pendientes por canal', 'ALTA', 'No', '--', 2000)
add('Policy Admin', s, p, 'GET', '/requirement/getpendingbytoken', 'Requerimientos pendientes por token', 'MEDIA', 'No', '--', 2000)
add('Policy Admin', s, p, 'POST', '/underwritingfiledocument', 'Subir documento de suscripcion', 'MEDIA', 'No', '--', 2000)
add('Policy Admin', s, p, 'GET', '/underwritingcase/getbyfilter', 'Casos de suscripcion filtrados', 'ALTA', 'No', '--', 2000)
add('Policy Admin', s, p, 'GET', '/Token/validate', 'Validar token de acceso unico', 'MEDIA', 'No', '--', 2000)

# --- Policy Admin - Risk ---
s, p = 'Risk', '/rest/risk/v1'
add('Policy Admin', s, p, 'GET', '/RiskInformation/opportunity/{OpportunityId}', 'Info de riesgo por oportunidad', 'ALTA', 'No', '--', 2000)
add('Policy Admin', s, p, 'GET', '/RiskInformation/quotation/{OfficialQuotation}', 'Info de riesgo por cotizacion', 'ALTA', 'No', '--', 2000)

# --- Policy Admin - BusinessRules ---
s, p = 'BusinessRules', '/rest/businessrules/v1'
add('Policy Admin', s, p, 'POST', '/Evaluate', 'Evaluar reglas de negocio', 'ALTA', 'No', '--', 2000)
add('Policy Admin', s, p, 'GET', '/Facts', 'Lista de hechos de reglas de negocio', 'MEDIA', 'No', '--', 2000)
add('Policy Admin', s, p, 'GET', '/Rulesets', 'Lista de paquetes de reglas disponibles', 'MEDIA', 'No', '--', 2000)

# --- Policy Admin - HealthDeclaration ---
s, p = 'HealthDeclaration', '/rest/healthdeclaration/v1'
add('Policy Admin', s, p, 'POST', '/healthdeclaration', 'Crear declaracion de salud', 'ALTA', 'No', '--', 2000)
add('Policy Admin', s, p, 'GET', '/healthdeclaration', 'Obtener declaracion por RecGUID', 'MEDIA', 'No', '--', 2000)
add('Policy Admin', s, p, 'GET', '/healthdeclarationpdf', 'PDF de declaracion de salud', 'BAJA', 'No', '--', 2000)

# --- Policy Admin - VehicleInformationService ---
s, p = 'VehicleInformationService', '/rest/vehicleinformationservice/v1'
add('Policy Admin', s, p, 'GET', '/VehicleModelByMake', 'Modelos de vehiculo por marca', 'MEDIA', 'No', '--', 2000)
add('Policy Admin', s, p, 'GET', '/VehicleEdition', 'Ediciones por modelo', 'MEDIA', 'No', '--', 2000)
add('Policy Admin', s, p, 'GET', '/VehicleValue', 'Valor del vehiculo', 'MEDIA', 'No', '--', 2000)

# --- Policy Admin - FleetService ---
add('Policy Admin', 'FleetService', '/rest/fleetservice/v1', 'POST', '/Vehicle', 'Registrar vehiculo en flota', 'MEDIA', 'No', '--', 2000)

# --- Policy Admin - Projection ---
add('Policy Admin', 'Projection', '/rest/projection/v1', 'POST', '/Request', 'Solicitar proyeccion de vida', 'MEDIA', 'No', '--', 2000)

# --- Policy Admin - ATS ---
s, p = 'ATS', '/rest/ats/v1'
add('Policy Admin', s, p, 'GET', '/quotation/byclient/{clientcode}', 'Cotizaciones por cliente (ATS)', 'MEDIA', 'No', '--', 2000)
add('Policy Admin', s, p, 'DELETE', '/AutomaticTesting', 'Eliminar datos de testing automatico', 'BAJA', 'No', '--', 2000)

# --- Facturacion y Cobranza - Billing ---
s, p = 'Billing', '/rest/billing/v1'
add('Facturacion y Cobranza', s, p, 'GET', '/Bill/BillsByPolicy', 'Recibos por poliza (PolicyID, EffectiveDate)', 'ALTA', 'Si', 'TG03', 2000)
add('Facturacion y Cobranza', s, p, 'GET', '/Bill/BillsByClient', 'Recibos por cliente (ClientID, EffectiveDate)', 'ALTA', 'Si', 'TG03', 2000)
add('Facturacion y Cobranza', s, p, 'POST', '/BillingOrder/BillingPolicyIssue', 'Orden de facturacion por emision de poliza', 'ALTA', 'No', '--', 2000)

# --- Facturacion y Cobranza - BusinessRules ---
s, p = 'BusinessRules', '/rest/businessrules/v1'
add('Facturacion y Cobranza', s, p, 'POST', '/Evaluate', 'Evaluar reglas (dias de gracia, mora)', 'ALTA', 'No', '--', 2000)
add('Facturacion y Cobranza', s, p, 'GET', '/Facts', 'Hechos de reglas de facturacion', 'MEDIA', 'No', '--', 2000)
add('Facturacion y Cobranza', s, p, 'GET', '/Rulesets', 'Paquetes de reglas disponibles', 'MEDIA', 'No', '--', 2000)

# --- Siniestros - FNOL ---
add('Siniestros', 'FNOL (NoticeOfLoss)', '/rest/fnol/v1', 'GET', '/BasicInformation', 'Informacion basica de siniestro (NoticeID)', 'ALTA', 'Si', 'TG02', 2000)

# --- Siniestros - BusinessRules ---
s, p = 'BusinessRules', '/rest/businessrules/v1'
add('Siniestros', s, p, 'POST', '/Evaluate', 'Evaluar reglas de validacion de cobertura', 'ALTA', 'No', '--', 2000)
add('Siniestros', s, p, 'GET', '/Facts', 'Hechos de reglas de siniestros', 'MEDIA', 'No', '--', 2000)
add('Siniestros', s, p, 'GET', '/Rulesets', 'Paquetes de reglas disponibles', 'MEDIA', 'No', '--', 2000)

# --- Distribution Management - DistributionChannelService ---
s, p = 'DistributionChannelService', '/rest/distributionchannelservice/v1'
# GET
add('Distribution Management', s, p, 'GET', '/DistributionChannel', 'Canal por ID', 'ALTA', 'Si', 'TG05', 1500)
add('Distribution Management', s, p, 'GET', '/DistributionChannel/GetBySyncUserId/{SyncUserId}', 'Canal por SyncUserId', 'ALTA', 'No', '--', 1500)
add('Distribution Management', s, p, 'GET', '/DistributionChannel/GetForDistributionChannel_AccessToPortal/{id}', 'Acceso portal del canal', 'MEDIA', 'No', '--', 1500)
add('Distribution Management', s, p, 'GET', '/ChannelCurrentAccount/GetByChannel/{id}', 'Cuenta corriente del canal', 'ALTA', 'No', '--', 1500)
add('Distribution Management', s, p, 'GET', '/ChannelAuthorizedUser/GetByChannel/{id}', 'Usuarios autorizados del canal', 'MEDIA', 'No', '--', 1500)
add('Distribution Management', s, p, 'GET', '/ChannelAuthorizedUser/GetUserSyncIdSeparatedByCommas/{id}', 'SyncIds de usuarios', 'BAJA', 'No', '--', 1500)
add('Distribution Management', s, p, 'GET', '/PhysicalAddress/GetByChannel/{id}', 'Direcciones del canal', 'MEDIA', 'No', '--', 1500)
add('Distribution Management', s, p, 'GET', '/PhysicalAddress/{DistributionChannelId}', 'Direccion especifica', 'MEDIA', 'No', '--', 1500)
add('Distribution Management', s, p, 'GET', '/Credentials/GetByChannel/{id}', 'Credenciales del canal', 'MEDIA', 'No', '--', 1500)
add('Distribution Management', s, p, 'GET', '/Credentials', 'Credenciales por email', 'MEDIA', 'No', '--', 1500)
add('Distribution Management', s, p, 'GET', '/DistributionChannelHistory/GetByChannel/{id}', 'Historial del canal', 'MEDIA', 'No', '--', 1500)
add('Distribution Management', s, p, 'GET', '/EconomicAgreements/GetByChannel/{id}', 'Acuerdos economicos', 'ALTA', 'No', '--', 1500)
add('Distribution Management', s, p, 'GET', '/Loans/GetByChannel/{id}', 'Prestamos del canal (por status)', 'ALTA', 'No', '--', 1500)
add('Distribution Management', s, p, 'GET', '/Loans/GetByChannelAll/{id}', 'Todos los prestamos', 'MEDIA', 'No', '--', 1500)
add('Distribution Management', s, p, 'GET', '/Loans/BalanceForChannel/{id}', 'Saldo de prestamos', 'ALTA', 'No', '--', 1500)
add('Distribution Management', s, p, 'GET', '/Loans/GetNewLoan/{id}', 'Nuevo prestamo', 'MEDIA', 'No', '--', 1500)
add('Distribution Management', s, p, 'GET', '/Loans/GetById/{LoanId}', 'Prestamo por ID', 'MEDIA', 'No', '--', 1500)
add('Distribution Management', s, p, 'GET', '/DistributionNetwork/GetByChannel/{id}', 'Red de distribucion', 'ALTA', 'No', '--', 1500)
add('Distribution Management', s, p, 'GET', '/RiskProfessionalPolicies/GetByChannel/{id}', 'Polizas de riesgo profesional', 'MEDIA', 'No', '--', 1500)
add('Distribution Management', s, p, 'GET', '/Support/GetByChannel/{id}', 'Soporte del canal', 'BAJA', 'No', '--', 1500)
add('Distribution Management', s, p, 'GET', '/Phone/GetByChannel/{id}', 'Telefonos del canal', 'MEDIA', 'No', '--', 1500)
add('Distribution Management', s, p, 'GET', '/Phone/{DistributionChannelId}', 'Telefono especifico', 'MEDIA', 'No', '--', 1500)
add('Distribution Management', s, p, 'GET', '/NetworkMember/GetByChannel/{id}', 'Miembros de red', 'ALTA', 'No', '--', 1500)
add('Distribution Management', s, p, 'GET', '/NetworkMember/GetChildrenOfAFather/{id}', 'Hijos en jerarquia', 'MEDIA', 'No', '--', 1500)
add('Distribution Management', s, p, 'GET', '/DistributionChannelContext/GetForDistributionChannel/{AuthUserSyncId}', 'Contexto del canal', 'MEDIA', 'No', '--', 1500)
add('Distribution Management', s, p, 'GET', '/DistributionChannelCreationRequest/{Param}/{Other}', 'Solicitud de creacion', 'MEDIA', 'No', '--', 1500)
add('Distribution Management', s, p, 'GET', '/ProofOfInsurance', 'Prueba de seguro por email', 'MEDIA', 'No', '--', 1500)
add('Distribution Management', s, p, 'GET', '/ProducerLicenseCategory', 'Categorias de licencia', 'BAJA', 'No', '--', 1500)
add('Distribution Management', s, p, 'GET', '/InsuranceCompany', 'Companias de seguros', 'BAJA', 'No', '--', 1500)
add('Distribution Management', s, p, 'GET', '/CredentialDoc', 'Documento de credencial', 'MEDIA', 'No', '--', 1500)
add('Distribution Management', s, p, 'GET', '/ProofOfInsuranceDoc', 'Documento prueba de seguro', 'MEDIA', 'No', '--', 1500)
add('Distribution Management', s, p, 'GET', '/DistributionChannelType', 'Tipos de canal', 'BAJA', 'No', '--', 1500)
add('Distribution Management', s, p, 'GET', '/AuthorizedUser', 'Usuarios autorizados por email', 'ALTA', 'No', '--', 1500)
add('Distribution Management', s, p, 'GET', '/AuthorizedUser/verificationlink/{Param}/{Other}', 'Enlace verificacion', 'BAJA', 'No', '--', 1500)
add('Distribution Management', s, p, 'GET', '/DistributionManagementSetting', 'Config de Distribution Mgmt', 'BAJA', 'No', '--', 1500)
add('Distribution Management', s, p, 'GET', '/SocialMedia', 'Redes sociales del canal', 'BAJA', 'No', '--', 1500)
add('Distribution Management', s, p, 'GET', '/TaskCountForDistributionChannel/{id}', 'Conteo de tareas', 'MEDIA', 'No', '--', 1500)
add('Distribution Management', s, p, 'GET', '/ContactPerson', 'Personas de contacto', 'MEDIA', 'No', '--', 1500)
add('Distribution Management', s, p, 'GET', '/comission', 'Comisiones por canal', 'ALTA', 'Si', 'TG05', 1500)
# POST
add('Distribution Management', s, p, 'POST', '/PhysicalAddress', 'Crear direccion', 'MEDIA', 'No', '--', 1500)
add('Distribution Management', s, p, 'POST', '/Credentials', 'Crear credenciales', 'ALTA', 'No', '--', 1500)
add('Distribution Management', s, p, 'POST', '/Loans/{UserSyncId}', 'Crear prestamo', 'ALTA', 'No', '--', 1500)
add('Distribution Management', s, p, 'POST', '/Loans/CurrencyAndInterestByTypeOfApproval/{id}', 'Moneda e interes por tipo', 'MEDIA', 'No', '--', 1500)
add('Distribution Management', s, p, 'POST', '/Loans/TemporalSave/{id}', 'Guardado temporal prestamo', 'MEDIA', 'No', '--', 1500)
add('Distribution Management', s, p, 'POST', '/Phone', 'Agregar telefono', 'MEDIA', 'No', '--', 1500)
add('Distribution Management', s, p, 'POST', '/DistributionChannelCreationRequest', 'Crear solicitud canal', 'ALTA', 'No', '--', 1500)
add('Distribution Management', s, p, 'POST', '/ProofOfInsurance', 'Crear prueba de seguro', 'MEDIA', 'No', '--', 1500)
add('Distribution Management', s, p, 'POST', '/CredentialDoc', 'Subir documento credencial', 'MEDIA', 'No', '--', 1500)
add('Distribution Management', s, p, 'POST', '/ProofOfInsuranceDoc', 'Subir doc prueba seguro', 'MEDIA', 'No', '--', 1500)
add('Distribution Management', s, p, 'POST', '/AuthorizedUser', 'Registrar usuario autorizado', 'ALTA', 'No', '--', 1500)
add('Distribution Management', s, p, 'POST', '/SocialMedia', 'Registrar red social', 'BAJA', 'No', '--', 1500)
add('Distribution Management', s, p, 'POST', '/ContactPerson', 'Registrar persona contacto', 'MEDIA', 'No', '--', 1500)
add('Distribution Management', s, p, 'POST', '/comission', 'Crear comision', 'ALTA', 'No', '--', 1500)
# PUT
add('Distribution Management', s, p, 'PUT', '/DistributionChannel', 'Actualizar canal', 'ALTA', 'No', '--', 1500)
add('Distribution Management', s, p, 'PUT', '/PhysicalAddress', 'Actualizar direccion', 'MEDIA', 'No', '--', 1500)
add('Distribution Management', s, p, 'PUT', '/Loans', 'Actualizar prestamo', 'MEDIA', 'No', '--', 1500)
add('Distribution Management', s, p, 'PUT', '/Phone', 'Actualizar telefono', 'MEDIA', 'No', '--', 1500)
add('Distribution Management', s, p, 'PUT', '/DistributionChannelCreationRequest/refresh/{P}/{O}', 'Reenviar invitacion', 'BAJA', 'No', '--', 1500)
add('Distribution Management', s, p, 'PUT', '/DistributionChannelCreationRequest/complete/{P}/{O}', 'Completar creacion canal', 'ALTA', 'No', '--', 1500)
add('Distribution Management', s, p, 'PUT', '/CredentialDoc/metadata', 'Actualizar metadata credencial', 'BAJA', 'No', '--', 1500)
add('Distribution Management', s, p, 'PUT', '/ProofOfInsuranceDoc/metadata', 'Actualizar metadata seguro', 'BAJA', 'No', '--', 1500)
add('Distribution Management', s, p, 'PUT', '/AuthorizedUser', 'Actualizar perfil usuario autorizado', 'MEDIA', 'No', '--', 1500)
add('Distribution Management', s, p, 'PUT', '/AuthorizedUser/verificationlink/refresh/{P}/{O}', 'Refrescar verificacion', 'BAJA', 'No', '--', 1500)
add('Distribution Management', s, p, 'PUT', '/AuthorizedUser/verificationlink/complete/{P}/{O}', 'Completar verificacion', 'BAJA', 'No', '--', 1500)
add('Distribution Management', s, p, 'PUT', '/SocialMedia', 'Actualizar red social', 'BAJA', 'No', '--', 1500)
add('Distribution Management', s, p, 'PUT', '/ContactPerson', 'Actualizar persona contacto', 'MEDIA', 'No', '--', 1500)
# DELETE
add('Distribution Management', s, p, 'DELETE', '/PhysicalAddress/{PhysicalAddressCode}', 'Eliminar direccion', 'BAJA', 'No', '--', 1500)
add('Distribution Management', s, p, 'DELETE', '/Credentials', 'Eliminar credencial', 'BAJA', 'No', '--', 1500)
add('Distribution Management', s, p, 'DELETE', '/Phone/{PhoneCode}', 'Eliminar telefono', 'BAJA', 'No', '--', 1500)
add('Distribution Management', s, p, 'DELETE', '/Loans/{UserSyncId}', 'Eliminar prestamo', 'BAJA', 'No', '--', 1500)
add('Distribution Management', s, p, 'DELETE', '/CredentialDoc/{CredentialFileID}', 'Eliminar doc credencial', 'BAJA', 'No', '--', 1500)
add('Distribution Management', s, p, 'DELETE', '/ProofOfInsuranceDoc/{ProofOfInsuranceFileID}', 'Eliminar doc seguro', 'BAJA', 'No', '--', 1500)
add('Distribution Management', s, p, 'DELETE', '/AuthorizedUser', 'Eliminar usuario autorizado', 'BAJA', 'No', '--', 1500)
add('Distribution Management', s, p, 'DELETE', '/SocialMedia', 'Eliminar red social', 'BAJA', 'No', '--', 1500)
add('Distribution Management', s, p, 'DELETE', '/ContactPerson/{ContactPersonId}', 'Eliminar contacto', 'BAJA', 'No', '--', 1500)

# --- Distribution Management - CommisionAccount ---
s, p = 'CommisionAccount', '/rest/commisionaccount/v1'
add('Distribution Management', s, p, 'GET', '/AccountPayableFromChannel', 'Cuentas por pagar del canal', 'ALTA', 'Si', 'TG05', 1500)
add('Distribution Management', s, p, 'POST', '/AccountPayableFromChannel', 'Crear cuenta por pagar', 'ALTA', 'No', '--', 1500)
add('Distribution Management', s, p, 'PUT', '/AccountPayableFromChannel', 'Actualizar cuenta por pagar', 'MEDIA', 'No', '--', 1500)
add('Distribution Management', s, p, 'DELETE', '/AccountPayableFromChannel', 'Eliminar cuenta por pagar', 'BAJA', 'No', '--', 1500)
add('Distribution Management', s, p, 'GET', '/AccountPayableFromChannel/AccountPayableID', 'Cuenta por ID', 'MEDIA', 'No', '--', 1500)
add('Distribution Management', s, p, 'GET', '/AccountPayableLine', 'Lineas de cuenta por pagar', 'ALTA', 'No', '--', 1500)
add('Distribution Management', s, p, 'POST', '/AccountPayableLine', 'Crear linea', 'MEDIA', 'No', '--', 1500)
add('Distribution Management', s, p, 'PUT', '/AccountPayableLine', 'Actualizar linea', 'MEDIA', 'No', '--', 1500)
add('Distribution Management', s, p, 'DELETE', '/AccountPayableLine', 'Eliminar linea', 'BAJA', 'No', '--', 1500)
add('Distribution Management', s, p, 'GET', '/CommissionLiquidation', 'Liquidacion de comisiones', 'ALTA', 'No', '--', 1500)
add('Distribution Management', s, p, 'PUT', '/CommissionLiquidation', 'Actualizar liquidacion', 'ALTA', 'No', '--', 1500)
add('Distribution Management', s, p, 'GET', '/CommissionLiquidation/ByStatus', 'Liquidaciones por estado', 'ALTA', 'No', '--', 1500)
add('Distribution Management', s, p, 'GET', '/CommissionLiquidation/ByChannel', 'Liquidaciones por canal', 'ALTA', 'No', '--', 1500)
add('Distribution Management', s, p, 'GET', '/AccountPayableAttachment', 'Obtener adjunto', 'MEDIA', 'No', '--', 1500)
add('Distribution Management', s, p, 'POST', '/AccountPayableAttachment', 'Subir adjunto', 'MEDIA', 'No', '--', 1500)
add('Distribution Management', s, p, 'PUT', '/AccountPayableAttachment/metadata', 'Actualizar metadata', 'BAJA', 'No', '--', 1500)
add('Distribution Management', s, p, 'DELETE', '/AccountPayableAttachment/{AttachmentFileId}', 'Eliminar adjunto', 'BAJA', 'No', '--', 1500)
add('Distribution Management', s, p, 'POST', '/AccountPayableHistory', 'Crear historial', 'MEDIA', 'No', '--', 1500)
add('Distribution Management', s, p, 'GET', '/AccountPayableHistory', 'Historial de cuenta', 'MEDIA', 'No', '--', 1500)
add('Distribution Management', s, p, 'POST', '/ChannelCurrentAccount', 'Crear cuenta corriente', 'ALTA', 'No', '--', 1500)
add('Distribution Management', s, p, 'GET', '/ChannelCurrentAccount', 'Cuentas corrientes', 'ALTA', 'No', '--', 1500)
add('Distribution Management', s, p, 'GET', '/EconomicAgreement', 'Acuerdo economico', 'ALTA', 'No', '--', 1500)
add('Distribution Management', s, p, 'GET', '/DetailByChannel', 'Detalle por canal', 'MEDIA', 'No', '--', 1500)

# --- Distribution Management - TrainingService ---
s, p = 'TrainingService', '/rest/trainingservice/v1'
add('Distribution Management', s, p, 'GET', '/Training/GetByUserSyncId/{UserId}/{ChannelId}', 'Entrenamientos por usuario y canal', 'MEDIA', 'No', '--', 1500)
add('Distribution Management', s, p, 'GET', '/Training', 'Entrenamientos por email', 'MEDIA', 'No', '--', 1500)
add('Distribution Management', s, p, 'POST', '/Training', 'Crear entrenamiento', 'MEDIA', 'No', '--', 1500)
add('Distribution Management', s, p, 'DELETE', '/Training', 'Eliminar entrenamiento', 'BAJA', 'No', '--', 1500)
add('Distribution Management', s, p, 'GET', '/InternalTraining/CurrentAll', 'Todos los entrenamientos internos', 'MEDIA', 'No', '--', 1500)
add('Distribution Management', s, p, 'GET', '/TrainingOption', 'Opciones de entrenamiento', 'BAJA', 'No', '--', 1500)
add('Distribution Management', s, p, 'GET', '/EducationInstitution', 'Instituciones educativas', 'BAJA', 'No', '--', 1500)

# --- Distribution Management - Integrations ---
s, p = 'Integrations', '/rest/integrations/v1'
add('Distribution Management', s, p, 'GET', '/commissionabletransaction', 'Listar transacciones comisionables', 'ALTA', 'No', '--', 1500)
add('Distribution Management', s, p, 'POST', '/commissionabletransaction', 'Crear transacciones comisionables', 'ALTA', 'No', '--', 1500)
add('Distribution Management', s, p, 'GET', '/commissionabletransaction/{transactionid}', 'Transaccion por ID', 'MEDIA', 'No', '--', 1500)
add('Distribution Management', s, p, 'DELETE', '/commissionabletransaction/{transactionid}', 'Eliminar transaccion', 'BAJA', 'No', '--', 1500)

# --- Ventas - OpportunityService ---
s, p = 'OpportunityService', '/rest/opportunityservice/v1'
add('Ventas', s, p, 'GET', '/Lead/{LeadId}', 'Obtener prospecto por ID', 'ALTA', 'No', '--', 1500)
add('Ventas', s, p, 'GET', '/Lead/GetAllByDistributionChannel/{ChannelSyncId}', 'Prospectos por canal', 'ALTA', 'No', '--', 1500)
add('Ventas', s, p, 'POST', '/Lead', 'Crear/actualizar oportunidad', 'ALTA', 'Si', 'TG04', 1500)
add('Ventas', s, p, 'GET', '/Lead', 'Prospecto por email', 'ALTA', 'Si', 'TG04', 1500)
add('Ventas', s, p, 'GET', '/Lead/CountAllByDistributionChannel/{id}/{OnlyLead}', 'Contar prospectos del canal', 'MEDIA', 'No', '--', 1500)
add('Ventas', s, p, 'GET', '/Opportunity/GetAllByDistributionChannel/{id}', 'Oportunidades por canal', 'ALTA', 'No', '--', 1500)
add('Ventas', s, p, 'GET', '/Opportunity/GetById/{OpportunityId}', 'Oportunidad por ID', 'ALTA', 'No', '--', 1500)
add('Ventas', s, p, 'GET', '/Opportunity/CountAllByDistributionChannel/{id}', 'Contar oportunidades del canal', 'MEDIA', 'No', '--', 1500)
add('Ventas', s, p, 'DELETE', '/Opportunity/all', 'Eliminar todas las oportunidades', 'BAJA', 'No', '--', 1500)
add('Ventas', s, p, 'GET', '/Opportunity/GetNewOpportunityForQuotation', 'Nueva oportunidad para cotizacion', 'ALTA', 'No', '--', 1500)
add('Ventas', s, p, 'GET', '/LeadAddress/All', 'Direcciones del prospecto', 'MEDIA', 'No', '--', 1500)
add('Ventas', s, p, 'POST', '/LeadAddress/Save', 'Guardar direccion', 'MEDIA', 'No', '--', 1500)
add('Ventas', s, p, 'GET', '/LeadAddress/RetrieveActiveByClient', 'Direcciones activas por cliente', 'MEDIA', 'No', '--', 1500)
add('Ventas', s, p, 'DELETE', '/LeadAddress', 'Eliminar direccion', 'BAJA', 'No', '--', 1500)
add('Ventas', s, p, 'GET', '/LeadAddress/RetrieveAllByClient', 'Todas las direcciones por cliente', 'MEDIA', 'No', '--', 1500)
add('Ventas', s, p, 'PUT', '/LeadAddress/Update', 'Actualizar direccion', 'MEDIA', 'No', '--', 1500)
add('Ventas', s, p, 'GET', '/LeadAddress/ValidateType', 'Validar tipo de direccion', 'BAJA', 'No', '--', 1500)
add('Ventas', s, p, 'GET', '/Client/GetByFilter', 'Buscar cliente (Email, ClientCode)', 'ALTA', 'No', '--', 1500)
add('Ventas', s, p, 'GET', '/Client/Exists', 'Verificar si existe cliente', 'MEDIA', 'No', '--', 1500)
add('Ventas', s, p, 'GET', '/ClientPhone/{LeadId}', 'Telefonos del prospecto', 'MEDIA', 'No', '--', 1500)
add('Ventas', s, p, 'POST', '/ClientPhone/{LeadId}', 'Agregar telefono', 'MEDIA', 'No', '--', 1500)
add('Ventas', s, p, 'PUT', '/ClientPhone/{LeadId}', 'Actualizar telefono', 'MEDIA', 'No', '--', 1500)
add('Ventas', s, p, 'DELETE', '/ClientPhone/{LeadId}', 'Eliminar telefono', 'BAJA', 'No', '--', 1500)

# --- Ventas - BrokerService ---
s, p = 'BrokerService', '/rest/brokerservice/v1'
add('Ventas', s, p, 'POST', '/lead', 'Crear prospecto', 'ALTA', 'No', '--', 1500)
add('Ventas', s, p, 'PUT', '/lead', 'Actualizar prospecto', 'ALTA', 'No', '--', 1500)
add('Ventas', s, p, 'DELETE', '/lead', 'Eliminar prospecto', 'BAJA', 'No', '--', 1500)
add('Ventas', s, p, 'POST', '/exclusiveleadrequest', 'Solicitud de prospecto exclusivo', 'ALTA', 'No', '--', 1500)
add('Ventas', s, p, 'GET', '/alloweddocumentperrole', 'Documentos permitidos por producto', 'MEDIA', 'No', '--', 1500)
add('Ventas', s, p, 'GET', '/alloweddocumentperrole/alloweddocumentbyallowedrole', 'Docs por rol', 'MEDIA', 'No', '--', 1500)
add('Ventas', s, p, 'GET', '/alloweddocumentperrole/clientdocumentbyopportunity', 'Doc de cliente por oportunidad', 'MEDIA', 'No', '--', 1500)
add('Ventas', s, p, 'GET', '/alloweddocumentperrole/clientdocument', 'Doc de cliente por ID', 'MEDIA', 'No', '--', 1500)
add('Ventas', s, p, 'GET', '/businessquoteinformation', 'Info de cotizaciones de negocio', 'ALTA', 'No', '--', 1500)
add('Ventas', s, p, 'GET', '/businessquoteinformation/TotalCount', 'Total cotizaciones', 'MEDIA', 'No', '--', 1500)
add('Ventas', s, p, 'POST', '/salessurchargediscount/PostSurchargeDiscount', 'Crear recargo/descuento/impuesto', 'ALTA', 'No', '--', 1500)
add('Ventas', s, p, 'GET', '/salessurchargediscount/getbyquotation', 'Recargo/descuento por cotizacion', 'MEDIA', 'No', '--', 1500)
add('Ventas', s, p, 'DELETE', '/salessurchargediscount', 'Eliminar recargo/descuento', 'BAJA', 'No', '--', 1500)
add('Ventas', s, p, 'GET', '/salessurchargediscount/alreadyexist', 'Validar duplicado', 'BAJA', 'No', '--', 1500)
add('Ventas', s, p, 'GET', '/leadhelpercontext', 'Contexto del prospecto', 'MEDIA', 'No', '--', 1500)

# --- Ventas - Campaign ---
s, p = 'Campaign', '/rest/campaign/v1'
add('Ventas', s, p, 'GET', '/Campaign/GetByUserSyncId/{UserId}/{ChannelId}', 'Campanas por usuario y canal', 'MEDIA', 'No', '--', 1500)
add('Ventas', s, p, 'GET', '/Campaign/GetByCampaignCodes', 'Campanas por codigos', 'MEDIA', 'No', '--', 1500)
add('Ventas', s, p, 'GET', '/Campaign', 'Campana por codigo', 'MEDIA', 'No', '--', 1500)

# --- Ventas - AuthenticationTest ---
add('Ventas', 'AuthenticationTest', '/rest/authenticationtest/v1', 'GET', '/Test', 'Health check de autenticacion', 'MEDIA', 'No', '--', 1500)

# --- Reaseguro ---
s, p = 'Reinsurance', '/rest/reinsurance/v1'
add('Reaseguro', s, p, 'GET', '/policytransactionnotification', 'Listar notificaciones de transaccion', 'ALTA', 'Si', 'TG08', 3000)
add('Reaseguro', s, p, 'POST', '/policytransactionnotification', 'Crear notificacion de transaccion', 'ALTA', 'Si', 'TG08', 3000)
add('Reaseguro', s, p, 'GET', '/policytransactionnotification/{id}', 'Notificacion por ID', 'ALTA', 'Si', 'TG09', 3000)
s, p = 'InsuranceCompanyService', '/rest/insurancecompanyservice/v1'
add('Reaseguro', s, p, 'GET', '/insurancecompany', 'Listar companias de seguros', 'ALTA', 'No', '--', 3000)
add('Reaseguro', s, p, 'GET', '/insurancecompany/{insurancecompanycode}', 'Compania por codigo', 'MEDIA', 'No', '--', 3000)
s, p = 'BusinessRules', '/rest/businessrules/v1'
add('Reaseguro', s, p, 'POST', '/Evaluate', 'Evaluar reglas de reaseguro', 'ALTA', 'Si', 'TG09', 3000)
add('Reaseguro', s, p, 'GET', '/Facts', 'Hechos de reglas de reaseguro', 'MEDIA', 'No', '--', 3000)
add('Reaseguro', s, p, 'GET', '/Rulesets', 'Paquetes de reglas disponibles', 'MEDIA', 'No', '--', 3000)

# --- Productos ---
s, p = 'ProductService', '/rest/product/v1'
add('Productos', s, p, 'GET', '/Product', 'Listar productos (paginado, filtros)', 'ALTA', 'Si', 'TG07', 3000)
add('Productos', s, p, 'GET', '/Product/image', 'Imagen del producto', 'BAJA', 'No', '--', 3000)
add('Productos', s, p, 'GET', '/Product/ProductSetting', 'Config del producto', 'ALTA', 'No', '--', 3000)
add('Productos', s, p, 'GET', '/Product/product', 'Producto por codigo', 'ALTA', 'Si', 'TG07', 3000)
add('Productos', s, p, 'GET', '/ExtraPremiumDiscountTax/All/{ProductCode}/{EffectiveDate}/{LOB}', 'Recargos/descuentos/impuestos', 'MEDIA', 'No', '--', 3000)
add('Productos', s, p, 'GET', '/Module', 'Modulos del producto', 'MEDIA', 'No', '--', 3000)
s, p = 'CoverageService', '/rest/coverageservice/v1'
add('Productos', s, p, 'GET', '/coverage/All/{ProductCode}/{LOB}/{EffectiveDate}', 'Coberturas por producto', 'ALTA', 'Si', 'TG07', 3000)
add('Productos', s, p, 'GET', '/coverageselectionhelper/GetByProductAndCoverageCode', 'Ayuda seleccion cobertura', 'MEDIA', 'No', '--', 3000)
s, p = 'LifeHealthService', '/rest/lifehealthservice/v1'
add('Productos', s, p, 'GET', '/AllowedCombinationOfPaymentAndInsuranceDuration/{code}/{lob}/{date}', 'Combinaciones pago/duracion', 'MEDIA', 'No', '--', 3000)
add('Productos', s, p, 'GET', '/AllowedCombinationOfPaymentAndInsuranceDuration/GetFirst/{code}/{lob}/{date}', 'Primera combinacion', 'MEDIA', 'No', '--', 3000)
add('Productos', 'PropertyAndCasualty', '/rest/propertyandcasualty/v1', 'GET', '/PropertyInsuredType/{ProductCode}/{LOB}/{EffectiveDate}', 'Tipos de asegurado P&C', 'MEDIA', 'No', '--', 3000)

# --- Admin Usuarios ---
add('Admin Usuarios', 'UserManagement', '/rest/usermanagement/v1', 'GET', '/user/image', 'Imagen de usuario', 'BAJA', 'Si', 'TG06', 3000)
s, p = 'UserManagementServices', '/rest/usermanagementservices/v1'
add('Admin Usuarios', s, p, 'POST', '/user', 'Crear usuario', 'ALTA', 'Si', 'TG06', 3000)
add('Admin Usuarios', s, p, 'PUT', '/user/activate', 'Activar cuenta de usuario (Token, SyncId)', 'ALTA', 'Si', 'TG11', 3000)
add('Admin Usuarios', s, p, 'PUT', '/user/Email', 'Cambiar email de usuario', 'ALTA', 'Si', 'TG11', 3000)

# --- SMC ---
s, p = 'SecureMessageCenter', '/rest/securemessagecenter/v1'
add('Servicio de Campo (SMC)', s, p, 'GET', '/SMC', 'Resumen de mensajes del usuario (SyncId)', 'MEDIA', 'Si', 'TG12', 3000)
add('Servicio de Campo (SMC)', s, p, 'POST', '/Email', 'Enviar correo centralizado', 'ALTA', 'Si', 'TG12', 3000)
add('Servicio de Campo (SMC)', 'TaskCountForUserService', '/rest/taskcountforuserservice/v1', 'GET', '/GetTaskCountForUser/{SyncId}', 'Conteo de tareas del usuario', 'MEDIA', 'No', '--', 3000)

# --- Portal Autoservicio ---
s, p = 'Data', '/rest/data/v1'
add('Portal Autoservicio', s, p, 'POST', '/validationmessage', 'Enviar mensajes de validacion', 'MEDIA', 'Si', 'TG10', 3000)
add('Portal Autoservicio', s, p, 'GET', '/validationmessage', 'Obtener mensajes de validacion', 'MEDIA', 'Si', 'TG10', 3000)

# --- Configuracion ---
add('Configuracion', 'SystemSetting', '/rest/systemsetting/v1', 'GET', '/generalsystemsetting', 'Configuracion general del sistema', 'ALTA', 'Si', 'TG14', 3000)
s, p = 'ExchangeRateService', '/rest/exchangerateservice/v1'
add('Configuracion', s, p, 'GET', '/ExchangeRate', 'Tipos de cambio (por SyncId)', 'MEDIA', 'No', '--', 3000)
add('Configuracion', s, p, 'GET', '/ExchangeRate/GetByCurrencyCode/{CurrencyCode}', 'Tipo cambio por moneda', 'MEDIA', 'No', '--', 3000)
s, p = 'BankService', '/rest/bankservice/v1'
add('Configuracion', s, p, 'GET', '/BankService/{BankCode}', 'Banco por codigo', 'MEDIA', 'No', '--', 3000)
add('Configuracion', s, p, 'GET', '/BankService/GetAll', 'Listar todos los bancos', 'MEDIA', 'No', '--', 3000)
add('Configuracion', s, p, 'GET', '/BankService/GetByCountry/{CountryCode}', 'Bancos por pais', 'MEDIA', 'No', '--', 3000)
add('Configuracion', s, p, 'GET', '/BankContactService/GetBySyncId/{SyncId}', 'Contacto bancario', 'BAJA', 'No', '--', 3000)
add('Configuracion', s, p, 'GET', '/BankContactService/GetFromBankCode/{BankCode}', 'Contactos del banco', 'BAJA', 'No', '--', 3000)
s, p = 'InterestService', '/rest/interestservice/v1'
add('Configuracion', s, p, 'GET', '/Interest', 'Tasas de interes (por SyncId)', 'MEDIA', 'No', '--', 3000)
add('Configuracion', s, p, 'GET', '/Interest/GetByCurrencyCode/{CurrencyCode}', 'Intereses por moneda', 'MEDIA', 'No', '--', 3000)
add('Configuracion', 'AccountPayableConceptService', '/rest/accountpayableconceptservice/v1', 'GET', '/AccountPayableConcept', 'Conceptos de cuentas por pagar', 'MEDIA', 'No', '--', 3000)

# --- OAuth (Transversal) ---
add('OAuth (Transversal)', 'OAuth', '/oauth/v2', 'POST', '/token', 'Obtener token OAuth (grant_type=password)', 'ALTA', 'No', '--', 1000)

print(f"Total endpoints: {len(endpoints)}")

# ============================================================
# STYLES
# ============================================================
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
FILL_HEADER = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
FONT_HEADER = Font(bold=True, color='FFFFFF', size=11)
FILL_GREEN = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
FILL_YELLOW = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
FILL_RED = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

METHOD_FILLS = {
    'GET': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
    'POST': PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid'),
    'PUT': PatternFill(start_color='F4B084', end_color='F4B084', fill_type='solid'),
    'DELETE': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
}
PRIO_FILLS = {
    'ALTA': PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid'),
    'MEDIA': PatternFill(start_color='F4B084', end_color='F4B084', fill_type='solid'),
    'BAJA': PatternFill(start_color='92D050', end_color='92D050', fill_type='solid'),
}
MODULE_COLORS = {
    'Policy Admin': PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid'),
    'Facturacion y Cobranza': PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid'),
    'Siniestros': PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid'),
    'Distribution Management': PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid'),
    'Ventas': PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
    'Reaseguro': PatternFill(start_color='E2D9F3', end_color='E2D9F3', fill_type='solid'),
    'Productos': PatternFill(start_color='D5F5E3', end_color='D5F5E3', fill_type='solid'),
    'Admin Usuarios': PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid'),
    'Servicio de Campo (SMC)': PatternFill(start_color='F9E79F', end_color='F9E79F', fill_type='solid'),
    'Portal Autoservicio': PatternFill(start_color='AED6F1', end_color='AED6F1', fill_type='solid'),
    'Configuracion': PatternFill(start_color='D5DBDB', end_color='D5DBDB', fill_type='solid'),
    'OAuth (Transversal)': PatternFill(start_color='F5CBA7', end_color='F5CBA7', fill_type='solid'),
}
SPRINT_FILLS = {
    'Sprint 1': PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid'),
    'Sprint 2': PatternFill(start_color='F4B084', end_color='F4B084', fill_type='solid'),
    'Sprint 3': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
    'Sprint 4': PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid'),
    'Sprint 5': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
}

mod_priority = {
    'Policy Admin': 'Alta', 'Facturacion y Cobranza': 'Alta', 'Siniestros': 'Alta',
    'Distribution Management': 'Alta', 'Ventas': 'Media', 'Reaseguro': 'Media',
    'Productos': 'Media', 'Admin Usuarios': 'Baja', 'Servicio de Campo (SMC)': 'Baja',
    'Portal Autoservicio': 'Baja', 'Configuracion': 'Baja', 'OAuth (Transversal)': 'Alta',
}
mod_order = ['Policy Admin', 'Facturacion y Cobranza', 'Siniestros', 'Distribution Management',
             'Ventas', 'Reaseguro', 'Productos', 'Admin Usuarios', 'Servicio de Campo (SMC)',
             'Portal Autoservicio', 'Configuracion', 'OAuth (Transversal)']

def style_header(ws, row, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

def auto_width(ws, min_w=10, max_w=50):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_w), max_w)

def assign_sprint(ep):
    m = ep['modulo']
    svc = ep['microservicio']
    if m == 'OAuth (Transversal)':
        return 'Sprint 1'
    if m == 'Policy Admin':
        return 'Sprint 1'
    if m in ('Siniestros', 'Facturacion y Cobranza'):
        return 'Sprint 2'
    if m == 'Distribution Management':
        return 'Sprint 3'
    if m in ('Ventas', 'Reaseguro'):
        return 'Sprint 4'
    return 'Sprint 5'

def estimate_hours(ep):
    return {'GET': 1, 'POST': 2, 'PUT': 1.5, 'DELETE': 1}.get(ep['metodo'], 1.5)

# ============================================================
# HOJA 1: Resumen Ejecutivo
# ============================================================
ws1 = wb.active
ws1.title = 'Resumen Ejecutivo'

ws1.merge_cells('A1:I1')
c = ws1['A1']
c.value = 'Control de Automatizacion de APIs - Policysense v1130'
c.font = Font(bold=True, size=16, color='1F4E79')
c.alignment = Alignment(horizontal='center', vertical='center')

total_auto = sum(1 for e in endpoints if e['automatizado'] == 'Si')
ws1.merge_cells('A2:I2')
c = ws1['A2']
c.value = f'Fecha: 2026-04-09 | Total: {len(endpoints)} endpoints | Cobertura: {total_auto}/{len(endpoints)} ({100*total_auto/len(endpoints):.1f}%)'
c.font = Font(size=11, italic=True)
c.alignment = Alignment(horizontal='center')

headers1 = ['Modulo', 'Microservicio(s)', 'Total Endpoints', 'Alta Prio', 'Media Prio', 'Baja Prio', 'Automatizados JMX', '% Cobertura', 'Prioridad Modulo']
for ci, h in enumerate(headers1, 1):
    ws1.cell(row=4, column=ci, value=h)
style_header(ws1, 4, len(headers1))

row = 5
totals = [0, 0, 0, 0, 0]
for mod in mod_order:
    mod_eps = [e for e in endpoints if e['modulo'] == mod]
    if not mod_eps:
        continue
    svcs = sorted(set(e['microservicio'] for e in mod_eps))
    total = len(mod_eps)
    alta = sum(1 for e in mod_eps if e['prioridad'] == 'ALTA')
    media = sum(1 for e in mod_eps if e['prioridad'] == 'MEDIA')
    baja = sum(1 for e in mod_eps if e['prioridad'] == 'BAJA')
    auto = sum(1 for e in mod_eps if e['automatizado'] == 'Si')
    pct = (auto / total * 100) if total > 0 else 0

    vals = [mod, ', '.join(svcs), total, alta, media, baja, auto, f'{pct:.1f}%', mod_priority.get(mod, 'Media')]
    for ci, v in enumerate(vals, 1):
        cell = ws1.cell(row=row, column=ci, value=v)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        if ci == 1:
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            if mod in MODULE_COLORS:
                cell.fill = MODULE_COLORS[mod]
        if ci == 2:
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    pct_cell = ws1.cell(row=row, column=8)
    if pct >= 70:
        pct_cell.fill = FILL_GREEN
    elif pct >= 40:
        pct_cell.fill = FILL_YELLOW
    else:
        pct_cell.fill = FILL_RED

    totals[0] += total; totals[1] += alta; totals[2] += media; totals[3] += baja; totals[4] += auto
    row += 1

total_pct = (totals[4] / totals[0] * 100) if totals[0] > 0 else 0
total_vals = ['TOTALES', '--', totals[0], totals[1], totals[2], totals[3], totals[4], f'{total_pct:.1f}%', '--']
for ci, v in enumerate(total_vals, 1):
    cell = ws1.cell(row=row, column=ci, value=v)
    cell.border = thin_border
    cell.font = Font(bold=True, size=11)
    cell.alignment = Alignment(horizontal='center', vertical='center')
tpc = ws1.cell(row=row, column=8)
tpc.fill = FILL_RED if total_pct < 40 else (FILL_YELLOW if total_pct < 70 else FILL_GREEN)

auto_width(ws1, 12, 55)
ws1.sheet_properties.tabColor = '1F4E79'

# ============================================================
# HOJA 2: Inventario Completo
# ============================================================
ws2 = wb.create_sheet('Inventario Completo')

headers2 = ['ID', 'Modulo', 'Microservicio', 'Servicio/Path base', 'Metodo HTTP', 'Endpoint',
            'Descripcion', 'Prioridad', 'Automatizado JMX', 'ThreadGroup JMX', 'SLA (ms)',
            'Estado Prueba', 'Notas']
for ci, h in enumerate(headers2, 1):
    ws2.cell(row=1, column=ci, value=h)
style_header(ws2, 1, len(headers2))

for i, ep in enumerate(endpoints, 2):
    vals = [ep['id'], ep['modulo'], ep['microservicio'], ep['path_base'], ep['metodo'],
            ep['endpoint'], ep['descripcion'], ep['prioridad'], ep['automatizado'],
            ep['tg'], ep['sla'], ep['estado'], ep['notas']]
    for ci, v in enumerate(vals, 1):
        cell = ws2.cell(row=i, column=ci, value=v)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        if ci == 2 and ep['modulo'] in MODULE_COLORS:
            cell.fill = MODULE_COLORS[ep['modulo']]
        if ci == 5:
            cell.fill = METHOD_FILLS.get(ep['metodo'], PatternFill())
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        if ci == 8:
            cell.fill = PRIO_FILLS.get(ep['prioridad'], PatternFill())
            cell.font = Font(bold=True, color='FFFFFF' if ep['prioridad'] == 'ALTA' else '000000')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        if ci == 9:
            cell.fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid') if ep['automatizado'] == 'Si' else PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        if ci == 12:
            if ep['estado'] == 'Completado':
                cell.fill = FILL_GREEN
            elif ep['estado'] == 'En progreso':
                cell.fill = FILL_YELLOW
            else:
                cell.fill = FILL_RED
            cell.alignment = Alignment(horizontal='center', vertical='center')

auto_width(ws2, 10, 55)
ws2.auto_filter.ref = f'A1:M{len(endpoints)+1}'
ws2.freeze_panes = 'A2'
ws2.sheet_properties.tabColor = '2E75B6'

# ============================================================
# HOJA 3: Plan de Automatizacion
# ============================================================
ws3 = wb.create_sheet('Plan de Automatizacion')

headers3 = ['ID', 'Modulo', 'Microservicio', 'Metodo', 'Endpoint', 'Descripcion',
            'Sprint Sugerido', 'Responsable', 'Estimacion (horas)']
for ci, h in enumerate(headers3, 1):
    ws3.cell(row=1, column=ci, value=h)
style_header(ws3, 1, len(headers3))

alta_not_auto = [e for e in endpoints if e['prioridad'] == 'ALTA' and e['automatizado'] == 'No']
alta_not_auto.sort(key=lambda x: (mod_order.index(x['modulo']) if x['modulo'] in mod_order else 99, x['id']))

row3 = 2
for ep in alta_not_auto:
    sprint = assign_sprint(ep)
    hrs = estimate_hours(ep)
    vals = [ep['id'], ep['modulo'], ep['microservicio'], ep['metodo'], ep['endpoint'],
            ep['descripcion'], sprint, 'Por asignar', hrs]
    for ci, v in enumerate(vals, 1):
        cell = ws3.cell(row=row3, column=ci, value=v)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        if ci == 2 and ep['modulo'] in MODULE_COLORS:
            cell.fill = MODULE_COLORS[ep['modulo']]
        if ci == 4:
            cell.fill = METHOD_FILLS.get(ep['metodo'], PatternFill())
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        if ci == 7:
            cell.fill = SPRINT_FILLS.get(sprint, PatternFill())
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
    row3 += 1

ws3.cell(row=row3+1, column=1, value='Total endpoints ALTA sin automatizar:').font = Font(bold=True, size=11)
ws3.cell(row=row3+1, column=5, value=len(alta_not_auto)).font = Font(bold=True, size=11, color='FF0000')
ws3.cell(row=row3+2, column=1, value='Estimacion total (horas):').font = Font(bold=True, size=11)
ws3.cell(row=row3+2, column=5, value=sum(estimate_hours(e) for e in alta_not_auto)).font = Font(bold=True, size=11, color='FF0000')

auto_width(ws3, 10, 55)
ws3.auto_filter.ref = f'A1:I{row3-1}'
ws3.freeze_panes = 'A2'
ws3.sheet_properties.tabColor = 'FF6B6B'

# ============================================================
# HOJA 4: Cobertura por Sprint
# ============================================================
ws4 = wb.create_sheet('Cobertura por Sprint')

ws4.merge_cells('A1:H1')
c = ws4['A1']
c.value = 'Plan de Cobertura por Sprint - Policysense Load Testing'
c.font = Font(bold=True, size=14, color='1F4E79')
c.alignment = Alignment(horizontal='center')

ws4.merge_cells('A2:H2')
c = ws4['A2']
c.value = 'Duracion estimada: 2 semanas por sprint | Solo endpoints ALTA prioridad sin automatizar'
c.font = Font(size=10, italic=True)
c.alignment = Alignment(horizontal='center')

sprint_defs = [
    ('Sprint 1', 'OAuth + Policy Admin (critico)', 'Autenticacion, cotizaciones, tarificacion, reglas de negocio', 'Semana 1-2'),
    ('Sprint 2', 'Siniestros + Facturacion (core)', 'FNOL, validacion cobertura, facturacion, reglas', 'Semana 3-4'),
    ('Sprint 3', 'Distribution Management', 'Canales, comisiones, liquidaciones, prestamos, redes', 'Semana 5-6'),
    ('Sprint 4', 'Ventas + Reaseguro', 'Oportunidades, prospectos, brokers, contratos reaseguro', 'Semana 7-8'),
    ('Sprint 5', 'Productos + UserMgmt + Portal + SMC + Settings', 'Catalogo productos, usuarios, portal, mensajeria, config', 'Semana 9-10'),
]

headers4 = ['Sprint', 'Foco', 'Modulos/Servicios', 'Periodo', 'Endpoints ALTA', 'Horas Estimadas', 'Acumulado %', 'Estado']
for ci, h in enumerate(headers4, 1):
    ws4.cell(row=4, column=ci, value=h)
style_header(ws4, 4, len(headers4))

total_alta_not_auto = len(alta_not_auto)
cumulative = 0
for i, (sname, foco, desc, periodo) in enumerate(sprint_defs, 5):
    sprint_eps = [e for e in alta_not_auto if assign_sprint(e) == sname]
    count = len(sprint_eps)
    hrs = sum(estimate_hours(e) for e in sprint_eps)
    cumulative += count
    pct_acum = (cumulative / total_alta_not_auto * 100) if total_alta_not_auto > 0 else 0

    vals = [sname, foco, desc, periodo, count, hrs, f'{pct_acum:.0f}%', 'Pendiente']
    for ci, v in enumerate(vals, 1):
        cell = ws4.cell(row=i, column=ci, value=v)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        if ci == 1:
            cell.fill = SPRINT_FILLS.get(sname, PatternFill())
            cell.font = Font(bold=True)
        if ci in (2, 3):
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

r = 11
ws4.cell(row=r, column=1, value='Resumen General').font = Font(bold=True, size=12, color='1F4E79')
ws4.cell(row=r+1, column=1, value='Total endpoints ALTA sin automatizar:')
ws4.cell(row=r+1, column=4, value=total_alta_not_auto).font = Font(bold=True)
ws4.cell(row=r+2, column=1, value='Total horas estimadas:')
ws4.cell(row=r+2, column=4, value=sum(estimate_hours(e) for e in alta_not_auto)).font = Font(bold=True)
ws4.cell(row=r+3, column=1, value='Endpoints ya automatizados:')
ws4.cell(row=r+3, column=4, value=total_auto).font = Font(bold=True)
ws4.cell(row=r+4, column=1, value='Cobertura actual:')
ws4.cell(row=r+4, column=4, value=f'{100*total_auto/len(endpoints):.1f}%').font = Font(bold=True, color='FF0000')
ws4.cell(row=r+5, column=1, value='Cobertura objetivo (5 sprints):')
ws4.cell(row=r+5, column=4, value='100% ALTA prioridad').font = Font(bold=True, color='00B050')

auto_width(ws4, 12, 55)
ws4.sheet_properties.tabColor = '00B050'

# ============================================================
# SAVE
# ============================================================
import os
output_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'docs', 'API_Control_Automatizacion.xlsx')
wb.save(output_path)
print(f"Guardado: {output_path}")
print(f"Hojas: {wb.sheetnames}")
print(f"Automatizados: {total_auto}")
print(f"ALTA sin automatizar: {len(alta_not_auto)}")
print("OK")
